import streamlit as st
import pandas as pd
import time
import zipfile
import base64
import tempfile
import openpyxl
from openpyxl.styles import Border, Side
from io import BytesIO





# Function to generate download link
def get_table_download_link(zip_filename):
    with open(zip_filename, 'rb') as f:
        b64 = base64.b64encode(f.read()).decode()
    href = f'<a href="data:file/zip;base64,{b64}" download="{zip_filename}">Download zip file</a>'
    return href

# Function to add borders to a sheet in an Excel workbook
def add_borders_to_sheet(sheet):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

# Button Execution Method
def button_execution(df, uploaded_file):
    progress_text = "报告生成中..."
    my_bar = st.progress(0, text=progress_text)
    
    #for i, file in enumerate(uploaded_file):
    for percent_complete in range(100):
        #time.sleep(0.01)
        df_sum = df.groupby('保（批）单号码')['实收保费'].sum().reset_index(name='实收保费汇总')
        df_merge = pd.merge(df, df_sum, how='left', on='保（批）单号码')
        df_merge = df_merge.drop_duplicates(subset='保（批）单号码')
        cols = df_merge.columns.tolist()
        b_index = cols.index('实收保费汇总')
        a_index = cols.index('实收保费')
        cols.insert(a_index, cols.pop(b_index))
        df_merge = df_merge[cols]
        df_merge = df_merge.drop(columns='实收保费')
                 
        
        my_bar.progress(percent_complete + 1)
    
    time.sleep(1)
        #my_bar.progress((i + 1) / len(uploaded_file),text=progress_text)

    with st.empty():
        st.write("后台正在将读取工作表")
        # Load the uploaded file into a workbook
        workbook_bytes = BytesIO(uploaded_file.getvalue())
        workbook = openpyxl.load_workbook(workbook_bytes)

        st.write('后台正在添加结果文件到工作表')
        # Add df_merge into this workbook
        with pd.ExcelWriter(workbook_bytes, engine='openpyxl') as writer:
            writer.book = workbook
            df_merge.to_excel(writer, sheet_name='公司_求和', index=False)
            writer.save()

        st.write('后台调整格式')
        # Add borders to the 'plus_sheet'
        workbook = openpyxl.load_workbook(workbook_bytes)
        sheet = workbook['公司_求和']
        add_borders_to_sheet(sheet)

        st.write('后台正在压缩文件')
        # Save the updated workbook to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            workbook.save(temp_file.name)
    
        # Zip the workbook
        zip_filename = '高杰手续费_求和.zip'
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            zipf.write(temp_file.name, arcname='高杰手续费_求和.xlsx')
 
        st.write('报告已生成,请点击下面按钮下载到本地📁')
    st.balloons()
    # Creating a download link
    st.markdown(get_table_download_link(zip_filename), unsafe_allow_html=True)




def main():
 
    # Add an image
    st.write('![Image](https://backiee.com/static/wpdb/wallpapers/v2/560x315/363086.jpg)')


    # Title
    st.write('###每月汇总###')
    
    # Sub-header
    st.subheader('请放入每月汇总文件xlsx')
    uploaded_file = st.file_uploader("请选择文件", accept_multiple_files=False, type=['xlsx'])


    if uploaded_file is not None:
        if '手续费' in uploaded_file.name:
            df = pd.read_excel(uploaded_file, sheet_name='公司')
            st.subheader('DataFrame')
            st.write(df)
            button = st.button('点击生产报告')
            if button:
                button_execution(df, uploaded_file)
        else:
            st.info('上传的文件名中没有包含"手续费"。')
    else:
        st.info('☝️ 请上传文件')
        
               


    #f button:
        #status = generate_reports(uploaded_files)
        #st.write(status)
        
    
if __name__ == "__main__":
    main()
